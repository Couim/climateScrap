from openpyxl import Workbook
import Constants
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import numbers 

#def toJSON(Months) :
    #toJson

##
#Function which take a list of months (Month object) and write it
#to an xlsx file
#@Months : list of months
##
def toXLSX(Months) :
    wb = Workbook()
    ws = wb.active
    rowNumber=2 #on démarre à la ligne 2 sous la légende
    columnNumber=1
    ws['B1']='jour'
    ws['C1']='temperature Min'
    ws['D1']='temperature Max'
    ws['E1']='pluie'
    # ws.merge_cells('B1:C1')
    ws.column_dimensions['B'].width=15
    thin_border_bottom = Border(bottom=Side(style='thin'))

    for month in Months:
        for day in month.days:
            ws[Constants.ALPHABET[columnNumber]+str(rowNumber)] = day.canonicForm
            ws[Constants.ALPHABET[columnNumber+1]+str(rowNumber)] = day.tempMin
            ws[Constants.ALPHABET[columnNumber+1]+str(rowNumber)].number_format = numbers.BUILTIN_FORMATS[4] 
            ws[Constants.ALPHABET[columnNumber+2]+str(rowNumber)] = day.tempMax
            ws[Constants.ALPHABET[columnNumber+2]+str(rowNumber)].number_format = numbers.BUILTIN_FORMATS[4]
            ws[Constants.ALPHABET[columnNumber+3]+str(rowNumber)] = day.pluie 
            ws[Constants.ALPHABET[columnNumber+3]+str(rowNumber)].number_format = numbers.BUILTIN_FORMATS[4]
            ws[Constants.ALPHABET[columnNumber]+str(rowNumber)].alignment = Alignment(horizontal='center')
            rowNumber = rowNumber+1

        ws[Constants.ALPHABET[columnNumber]+str(rowNumber-1)].border = thin_border_bottom
        ws[Constants.ALPHABET[columnNumber+1]+str(rowNumber-1)].border = thin_border_bottom
        ws[Constants.ALPHABET[columnNumber+2]+str(rowNumber-1)].border = thin_border_bottom
        ws[Constants.ALPHABET[columnNumber+3]+str(rowNumber-1)].border = thin_border_bottom
    wb.save('SMH2019_infoclimat.xlsx')
    print('Le fichier SMH2019_infoclimat.xslx a été écrit')


#def toXML(Months):
    #toXML